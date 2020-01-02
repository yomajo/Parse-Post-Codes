import os
import sys
import logging
import re
from textwrap import wrap
import openpyxl

#LOGGING SETUP
log_path = os.path.join(os.path.dirname(__file__), 'txt_to_excel.log') 
logging.basicConfig(level='INFO', filename=log_path, filemode='w')

# GLOBAL VARIABLES
txt_file = 'post_codes.txt' #Default value
split_at_chars = 13
xlsx_file = 'Postal_Codes_Manager.xlsx'
sheet_names = ['Free Codes', 'Expired Codes']
re_pattern = re.compile(r'UA[0-9]{9}LT')

# PATH BUILDING
wb_path = os.path.join(os.path.dirname(__file__), xlsx_file)
# If sys argument with path has been provided - use it as source abs path of text file, if not, default to hardcoded name.
if len(sys.argv) > 1:
    txt_path = sys.argv[1]
    logging.info(f'This path was given through system arguments: {txt_path}')
else:
    txt_path = os.path.join(os.path.dirname(__file__), txt_file)
    logging.info(f'No system arguments were passed; defaulting to hardcoded path: {txt_path}')


def convert_txt_to_list(txt_path):
    '''takes argument of txt file with a block of codes without spaces like 'UA982420356LTUA522331316LTUA675574744LT'
    and returns a list'''
    logging.info(f'Opening file to read: {txt_path}')
    with open(txt_path, 'r') as f:
        codes_block = f.read()
    # Code block to list:
    return wrap(codes_block, split_at_chars)

def edit_wb(xlsx_path):
    '''takes xlsx path, edits and saves workbook'''
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_names[0]]
    last_row = ws.max_row
    logging.info(f'Filling workbook with codes')
    add_codes(ws, last_row, 1)
    # NEW LINE BELOW
    logging.info(f'Currently workbook contains {ws.max_row} codes')
    # NEW LINE ABOVE
    logging.info(f'Workbook {xlsx_file} updated with new codes')
    wb.save(xlsx_path)
    wb.close()

def add_codes(ws, last_row, fill_col):
    '''iterates simultaneously through list and cells to transfer postal codes to xlsx'''
    codes_list = convert_txt_to_list(txt_path)
    validated_codes = validate_codes(codes_list)
    logging.info(f'Read {txt_file} and formed a list codes_list containing {len(validated_codes)} codes')
    # Handling the case of completely empty FreeCodes sheet, so filling would start at 1 row instead of 2
    if last_row == 1:
        last_row = 0
    for idx in range(1, len(validated_codes)+1):
        # Setting values:
        ws.cell(row=last_row + idx, column=fill_col).value = validated_codes[idx-1]
    logging.info('Finished looping through codes')

def validate_codes(unvalidated_list):
    '''Create a new list of passed one, filtering members through regex pattern'''
    validated_codes = []
    logging.info(f'Before validation, list len: {len(unvalidated_list)}')
    for code in unvalidated_list:
        if bool(re.match(re_pattern, code)):
            validated_codes.append(code)
    logging.info(f'Before validation, list len: {len(validated_codes)}')
    return validated_codes

def get_last_used_row_number(xlsx_path):
    '''takes xlsx file, opens, reads last row number and returns'''
    logging.info(f'Opening: {xlsx_path} to get last used row inside get_last_used_row_number func')
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Free Codes']
    last_row = ws.max_row
    wb.close()
    return last_row

def run():
    '''Main sequence of functions when executing script'''
    print(f'Logging to file: {log_path}')
    logging.info('Run() start')
    last_row_before = get_last_used_row_number(wb_path)
    edit_wb(wb_path)
    logging.info(f'SCRIPT FINISHED WITHOUT ERRORS')
    last_row_after = get_last_used_row_number(wb_path)
    print(f'SCRIPT FINISHED WITHOUT ERRORS')
    # Output to cmd for VBA to capture output:
    if last_row_after - last_row_before > 0:
        print('LOADED')
    else:
        print('FAILED')

if __name__ == "__main__":
    run()